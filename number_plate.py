import cv2
import pytesseract
from openpyxl import load_workbook, Workbook

# Load existing Excel workbook and sheet
try:
    wb = load_workbook("plates_info.xlsx")
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    ws.append(["Plate Number", "Slot", "X", "Y", "Width", "Height"])

harcascade = "model/indian_license_plate.xml"

cap = cv2.VideoCapture(0)

cap.set(3, 640) # width
cap.set(4, 480) # height

min_area = 500
count = 0
slots = 1  # Initialize slots counter

while True:
    success, img = cap.read()

    plate_cascade = cv2.CascadeClassifier(harcascade)
    img_gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    plates = plate_cascade.detectMultiScale(img_gray, 1.1, 4)

    largest_plate = None
    largest_area = 0

    for (x, y, w, h) in plates:
        area = w * h

        if area > largest_area:
            largest_area = area
            largest_plate = (x, y, w, h)

    if largest_plate is not None:
        x, y, w, h = largest_plate
        cv2.rectangle(img, (x, y), (x+w, y+h), (0, 255, 0), 2)
        cv2.putText(img, "Number Plate", (x, y-5), cv2.FONT_HERSHEY_COMPLEX_SMALL, 1, (255, 0, 255), 2)

        img_roi = img[y: y+h, x:x+w]
        cv2.imshow("ROI", img_roi)

        # Save plate image when 's' is pressed
        key = cv2.waitKey(1)
        if key == ord('s'):
            cv2.imwrite(f"plates/scaned_img_{count}.jpg", img_roi)
            cv2.rectangle(img, (0, 200), (640, 300), (0, 255, 0), cv2.FILLED)
            cv2.putText(img, "Plate Saved", (150, 265), cv2.FONT_HERSHEY_COMPLEX_SMALL, 2, (0, 0, 255), 2)
            cv2.imshow("Result", img)
            cv2.waitKey(500)
            count += 1

            # Perform OCR on saved plate image
            saved_img = cv2.imread(f"plates/scaned_img_{count - 1}.jpg")
            saved_img_gray = cv2.cvtColor(saved_img, cv2.COLOR_BGR2GRAY)
            saved_plate_text = pytesseract.image_to_string(saved_img_gray, config='--psm 8')
            saved_plate_text = saved_plate_text.replace('\n', '')  # Remove newline characters

            # Save plate information and OCR result to Excel sheet
            ws.append([saved_plate_text, slots, x, y, w, h])
            wb.save("plates_info.xlsx")

            # Increment slots counter
            slots += 1

    cv2.imshow("Result", img)

    key = cv2.waitKey(1)

    if key == ord('q'):  # Press 'q' to close the window
        break

cap.release()
cv2.destroyAllWindows()
